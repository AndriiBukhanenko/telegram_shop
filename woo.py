from woocommerce import API
import time

# from urllib.parse import urlencode
#
# store_url = 'http://wp.tigzver.ru.xsph.ru/'
# endpoint = '/wc-auth/v1/authorize'
# params = {
#     "app_name": "My App Name",
#     "scope": "read_write",
#     "user_id": 123,
#     "return_url": "http://app.com/return-page",
#     "callback_url": "https://app.com/callback-endpoint"
# }
# query_string = urlencode(params)
#
# print("%s%s?%s" % (store_url, endpoint, query_string))

wcapi = API(
    url="http://wp.tigzver.ru.xsph.ru/",
    consumer_key="ck_dfc4d91ffa849991dd1dc72a150bdde13d447f96",
    consumer_secret="cs_a1159c7cd22a6e66ca4b1e72c4c41392075a6e48",
    query_string_auth=True,
    version="wc/v3"
)





categories = []


def categories_brand():
    parfum = ['Atelier Cologne',
              'Atelier Des Ors',
              "Etat Libre d'Orange",
              'Evody Parfums',
              'Histoires de Parfums',
              'Hugh Parsons',
              'I love New York by Bond №9',
              "I Profumi di D'Annunzio",
              'Jeroboam',
              'Jovoy',
              'JUSBOX perfumes',
              'Laboratorio Olfattivo',
              'Lucien Ferrero Maître Parfumeur',
              'Miller et Bertaux',
              'Moresque Parfum',
              'Olibere',
              'Paglieri 1876',
              'Panama 1924',
              'Parfums De Marly',
              'Parfums d’Orsay',
              'Parfums Houbigant Paris',
              'Perris Monte Carlo',
              'Pineider',
              'Teo Cabanel',
              'The Different Company',
              'The House Of Oud',
              'The Merchant of Venice',
              'Wide Society']

    aromat = ['Jovoy',
              'Laboratorio Olfattivo',
              'Miller et Bertaux',
              'Moresque Parfum',
              'Parfums De Marly',
              'Pineider',
              'The Different Company']

    uxod = ['4711',
            'Erborian',
            'Hugh Parsons',
            'Laboratorio Olfattivo',
            'Panama 1924',
            'Parfums De Marly',
            'Parfums Houbigant Paris']

    data = [({"name": 'Парфюмерия'}, parfum),
            ({"name": 'Ароматы для дома'}, aromat),
            ({"name": 'Уход'}, uxod)]
    for item in data:
        id = wcapi.post("products/categories", item[0]).json()
        categories.append(id)
        print(id)
        time.sleep(1)
        for el in item[1]:
            data = {
                "parent": id['id'],
                "name": el
            }
            par = wcapi.post("products/categories", data).json()
            categories.append(par)




attributes_all = []

def attributes():
    vid = ['Духи',
           'Колонь абсолю',
           'Набор',
           'Парфюмированная вода',
           'Туалетная вода',
           'Аромат для дома',
           'Свеча',
           'PP Крем (Праймер)',
           'Дезодорант-стик',
           'Крем для тела',
           'Лосьон для тела',
           'Лосьон после бритья',
           'Мыло',
           'Уход за кожей глаз',
           'Увлажнение и матирование',
           'Тонизирование кожи',
           'СС крем (Сontrol Сolor)',
           'Скрабы и маски',
           'Очищение кожи',
           'Идеальные губы',
           'Восстановление кожи',
           'ВВ крем (Blemish Balm)',
           'Антивозврастной уход',
           '0.09 Линия',
           'Совершенствование и уход',
           'Одеколон',
           'Гель для душа']

    obem = ['125',
            '15',
            '15X3',
            '32',
            '75',
            '4'
            '78',
            '7x2',
            '30',
            '50',
            '60',
            '90',
            '100',
            '120',
            '200']

    aromat = ['Абсолютные',
              'Ароматные',
              'Гесперидные',
              'Гурманский',
              'Синтетические',
              'Альдегидные',
              'Амбровые',
              'Восточные',
              'Древесные',
              'Зелные',
              'Кожаные',
              'Мускусные',
              'Озоновый',
              'Пряные',
              'Пудровый',
              'Свежие',
              'Удовые',
              'Фруктовые',
              'Фужерные',
              'Цветочные',
              'Цитрусовые',
              'Шипровые',
              'Ягодные',
              'Морские']

    parfumer = ['Aglaé Nicolas',
                'Alexandra Kosinski',
                'Alexandra Monet',
                'Amelie Bourgeois',
                'Andrea Thero',
                'Andrea Thero Casotti',
                'Anne Sophie Behaghel',
                'Antoine Lie',
                'Antoine Maisondieu',
                'Aurelien Guichard',
                'Benoist Lapouza',
                'Bertrand Duchaufour',
                'Bruno Jovanovic',
                'Calice Becker',
                'Caroline Sabas',
                'Carlo Ribero',
                'Cecile Matton Polge',
                'Cecile Zarokian',
                'Celine Ellena',
                'Christine Nagel',
                'Corinne Cachen',
                'Cristian Calabrò',
                'David Maruitte',
                'Daniela Andrier',
                'Dominique Ropion',
                'Emilie Coppermann',
                'Enrico Buccella',
                'Fabrice Pellegrin',
                'Francis Deleamont',
                'Gerald Ghislain',
                'Gian Luca Perris',
                'Guillaume Flavigny',
                'Hamid Merati-Kashani',
                'Isabelle Fritsch',
                'Jacques Flori',
                'Jean Claude Ellena',
                'Jean-Francois Latty',
                'Jerome Epinette',
                'Julien Rasquinet',
                'Laurice Rahme',
                'Lucien Ferrero',
                'Luca Maffei',
                "Marc Fantond' Andon",
                'Marie Duchene',
                'Marie- Anne de Puyraimond',
                'Marie Schnirer',
                'Marie Salamagne',
                'Mathieu Nardin',
                'Mathilde Bijaoui',
                'Maurizio Cerizza',
                'Michele Boellis',
                'Michel Girard',
                'Michele Saramito',
                'Michelle Saramitot',
                'Nanako Ogi',
                'Nathalie Feisthauer',
                'Natalie Gracia Cetto',
                'Nathalie Lorson',
                'Olivier Pescheux',
                'Paul Parquet',
                'Philippe Romano',
                'Quentin Bisch',
                'Ralf Schweieger',
                'Robert Bienaime',
                'Shyamala Maisondieu',
                'Sidonie Lancesseur',
                'Vanina Murraciole',
                'Violaine Collas',
                'Yasnn Vasnier']

    notu = ['Абельмош',
            'Абрикос',
            'Абсент',
            'Агаровое дерево',
            'Адреналин',
            'Акация',
            'Акигалавуд',
            'Альдегиды',
            'Альдегид',
            'Амбра',
            'Амброксан',
            'Ананас',
            'Ангелика',
            'Анис',
            'Апельсин',
            'Арбуз',
            'Артемизия',
            'Базилик',
            'Бальзам',
            'Бальзамическая ель',
            'Банан',
            'Барриковое дерево',
            'Бархатцы',
            'Белые цветы',
            'Бензоин',
            'Бергамот',
            'Береза',
            'Бессмертник',
            'Бобы тонка',
            'Боярышник',
            'Ваниль',
            'Вербена',
            'Вереск',
            'Ветивер',
            'Винил',
            'Виола',
            'Виски',
            'Вишня',
            'Гальбанум',
            'Гардения',
            'Гваяковое дерево',
            'Гвоздика',
            'Гелиотроп',
            'Герань',
            'Гесперидий',
            'Гиацинт',
            'Гирециум',
            'Грейпфрут',
            'Груша',
            'Давана',
            'Дамасская роза',
            'Дерево',
            'Дерево Хиноки',
            'Джин',
            'Дуб',
            'Дубовый мох',
            'Дым',
            'Ежевика',
            'Еловый бальзам',
            'Жасмин',
            'Жимолость',
            'Желтые цветы',
            'Замша',
            'Зеленые листья',
            'Зефир',
            'Иланг-иланг',
            'Имбирь',
            'Инжир',
            'Ирис',
            'Иссоп',
            'Какао',
            'Камень',
            'Карамель',
            'Кардамон',
            'Карри',
            'Кастореум',
            'Кашемир',
            'Кашмеран',
            'Каштан',
            'Кедр',
            'Кипарис',
            'Клементин',
            'Клубника',
            'Кожа',
            'Кокос',
            'Кориандр',
            'Корица',
            'Космон',
            'Костус',
            'Кофе',
            'Кровь',
            'Кумарин',
            'Кумин',
            'Кунжут',
            'Куркума',
            'Крем Шантильи',
            'Лабданум',
            'Лаванда',
            'Лавандин',
            'Лавровые листья',
            'Ладан',
            'Ладанник',
            'Лайм',
            'Ландыш',
            'Лён',
            'Лилия',
            'Лимон',
            'Лимонник',
            'Листья Айвы',
            'Листья камелии',
            'Листья фиалки',
            'Личи',
            'Лишайник',
            'Лоренокс',
            'Лотос',
            'Любисток',
            'Магнолия',
            'Майоран',
            'Малина',
            'Манго',
            'Мандарин',
            'Маракуйя',
            'Маргаритка',
            'Марципан',
            'Масала',
            'Мате',
            'Мёд',
            'Ментол',
            'Мимоза',
            'Миндаль',
            'Мирра',
            'Мирт',
            'Можжевельник',
            'Молоко',
            'Морковь',
            'Морские ноты',
            'Мох',
            'Мускатный орех',
            'Мускус',
            'Мята',
            'Нарцисс',
            'Нектарин',
            'Нероли',
            'Одуванчик',
            'Озон',
            'Олибанум',
            'Олива',
            'Опопонакс',
            'Орхидея',
            'Османтус',
            'Осока папируса',
            'Пало Санто',
            'Папирус',
            'Пачули',
            'Пралине',
            'Перец',
            'Персик',
            'Петалия',
            'Петитгрейн',
            'Пион',
            'Плющ',
            'Подсолнух',
            'Полынь горькая',
            'Помело',
            'Попкорн',
            'Пряности',
            'Ревень',
            'Рис басмати',
            'Роза',
            'Розмарин',
            'Розовый перец',
            'Ром',
            'Ромашка',
            'Серебристая',
            'Сандаловое дерево',
            'Сандал',
            'Сандал Майсур',
            'Семена моркови',
            'Сахарный тростник',
            'Сандал Амириса',
            'Сено',
            'Сирень',
            'Слива',
            'Смоква',
            'Смола',
            'Смородина',
            'Солодка',
            'Сосна',
            'Стиракс',
            'Сухофрукты',
            'Табак',
            'Тиаре',
            'Тимьян',
            'Тмин',
            'Трава',
            'Трюфель',
            'Тубероза',
            'Уд',
            'Фиалка',
            'Фиговое дерево',
            'Финик',
            'Франжипани',
            'Фрезия',
            'Фрукты',
            'Хлопок',
            'Цветы апельсина',
            'Цветы груши',
            'Цветы малины',
            'Цикламен',
            'Циприол',
            'Цитрон',
            'Цитрус',
            'Чай',
            'Черника',
            'Чернослив',
            'Чили',
            'Эвернил',
            'Эвкалипт',
            'Элеми',
            'Эмблика',
            'Эстрагон',
            'Эфир',
            'Шалфей',
            'Шампанское',
            'Шафран',
            'Шелковое дерево',
            'Шерсть',
            'Шиповник',
            'Шоколад',
            'Яблоко',
            'Ягоды',
            'Ягоды годжи',
            'Гранат',
            'Клюква',
            'Пасифлора']

    attributes_list = [['Вид', vid], ['Ароматы', aromat], ['Ноты',notu], ['Обьём', obem], ['Парфюмер', parfumer]]
    for atr in attributes_list:
        data = {
            "name": atr[0],
            "type": "select",
            "order_by": "menu_order",
            "has_archives": True
        }
        id = wcapi.post("products/attributes", data).json()
        print(id)
        attributes_all.append(id)
        for el in atr[1]:
            data = {
                "name": el
            }
            wcapi.post(f"products/attributes/{id['id']}/terms", data).json()


# categories_brand()
# attributes()


from openpyxl import load_workbook

categories = [{'id': 1153, 'name': 'Парфюмерия', 'slug': '%d0%bf%d0%b0%d1%80%d1%84%d1%8e%d0%bc%d0%b5%d1%80%d0%b8%d1%8f', 'parent': 0, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}]}}, {'id': 1154, 'name': 'Atelier Cologne', 'slug': 'atelier-cologne', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1154'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1155, 'name': 'Atelier Des Ors', 'slug': 'atelier-des-ors', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1155'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1156, 'name': "Etat Libre d'Orange", 'slug': 'etat-libre-dorange', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1156'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1157, 'name': 'Evody Parfums', 'slug': 'evody-parfums', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1157'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1158, 'name': 'Histoires de Parfums', 'slug': 'histoires-de-parfums', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1158'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1159, 'name': 'Hugh Parsons', 'slug': 'hugh-parsons', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1159'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1160, 'name': 'I love New York by Bond №9', 'slug': 'i-love-new-york-by-bond-%e2%84%969', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1160'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1161, 'name': "I Profumi di D'Annunzio", 'slug': 'i-profumi-di-dannunzio', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1161'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1162, 'name': 'Jeroboam', 'slug': 'jeroboam', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1162'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1163, 'name': 'Jovoy', 'slug': 'jovoy', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1163'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1164, 'name': 'JUSBOX perfumes', 'slug': 'jusbox-perfumes', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1164'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1165, 'name': 'Laboratorio Olfattivo', 'slug': 'laboratorio-olfattivo', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1165'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1166, 'name': 'Lucien Ferrero Maître Parfumeur', 'slug': 'lucien-ferrero-maitre-parfumeur', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1166'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1167, 'name': 'Miller et Bertaux', 'slug': 'miller-et-bertaux', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1167'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1168, 'name': 'Moresque Parfum', 'slug': 'moresque-parfum', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1168'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1169, 'name': 'Olibere', 'slug': 'olibere', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1169'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1170, 'name': 'Paglieri 1876', 'slug': 'paglieri-1876', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1170'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1171, 'name': 'Panama 1924', 'slug': 'panama-1924', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1171'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1172, 'name': 'Parfums De Marly', 'slug': 'parfums-de-marly', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1172'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1173, 'name': 'Parfums d’Orsay', 'slug': 'parfums-dorsay', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1173'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1174, 'name': 'Parfums Houbigant Paris', 'slug': 'parfums-houbigant-paris', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1174'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1175, 'name': 'Perris Monte Carlo', 'slug': 'perris-monte-carlo', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1175'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1176, 'name': 'Pineider', 'slug': 'pineider', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1176'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1177, 'name': 'Teo Cabanel', 'slug': 'teo-cabanel', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1177'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1178, 'name': 'The Different Company', 'slug': 'the-different-company', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1178'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1179, 'name': 'The House Of Oud', 'slug': 'the-house-of-oud', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1179'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1180, 'name': 'The Merchant of Venice', 'slug': 'the-merchant-of-venice', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1180'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1181, 'name': 'Wide Society', 'slug': 'wide-society', 'parent': 1153, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1181'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1153'}]}}, {'id': 1182, 'name': 'Ароматы для дома', 'slug': '%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 0, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}]}}, {'id': 1183, 'name': 'Jovoy', 'slug': 'jovoy-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1183'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1184, 'name': 'Laboratorio Olfattivo', 'slug': 'laboratorio-olfattivo-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1184'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1185, 'name': 'Miller et Bertaux', 'slug': 'miller-et-bertaux-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1185'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1186, 'name': 'Moresque Parfum', 'slug': 'moresque-parfum-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1186'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1187, 'name': 'Parfums De Marly', 'slug': 'parfums-de-marly-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1187'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1188, 'name': 'Pineider', 'slug': 'pineider-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1188'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1189, 'name': 'The Different Company', 'slug': 'the-different-company-%d0%b0%d1%80%d0%be%d0%bc%d0%b0%d1%82%d1%8b-%d0%b4%d0%bb%d1%8f-%d0%b4%d0%be%d0%bc%d0%b0', 'parent': 1182, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1189'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1182'}]}}, {'id': 1190, 'name': 'Уход', 'slug': '%d1%83%d1%85%d0%be%d0%b4', 'parent': 0, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}]}}, {'id': 1191, 'name': '4711', 'slug': '4711', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1191'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}, {'id': 1192, 'name': 'Erborian', 'slug': 'erborian', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1192'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}, {'id': 1193, 'name': 'Hugh Parsons', 'slug': 'hugh-parsons-%d1%83%d1%85%d0%be%d0%b4', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1193'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}, {'id': 1194, 'name': 'Laboratorio Olfattivo', 'slug': 'laboratorio-olfattivo-%d1%83%d1%85%d0%be%d0%b4', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1194'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}, {'id': 1195, 'name': 'Panama 1924', 'slug': 'panama-1924-%d1%83%d1%85%d0%be%d0%b4', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1195'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}, {'id': 1196, 'name': 'Parfums De Marly', 'slug': 'parfums-de-marly-%d1%83%d1%85%d0%be%d0%b4', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1196'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}, {'id': 1197, 'name': 'Parfums Houbigant Paris', 'slug': 'parfums-houbigant-paris-%d1%83%d1%85%d0%be%d0%b4', 'parent': 1190, 'description': '', 'display': 'default', 'image': None, 'menu_order': 0, 'count': 0, '_links': {'self': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1197'}], 'collection': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories'}], 'up': [{'href': 'http://wp.tigzver.ru.xsph.ru/wp-json/wc/v3/products/categories/1190'}]}}]

def push_products():

    wb = load_workbook('./products.xlsx')

    sheet = wb.get_sheet_by_name('Лист1')

    for i in range(2, 25):
         print(i, sheet.cell(row=i, column=2).value)
         name = sheet.cell(row=i, column=3).value
         price = sheet.cell(row=i, column=7).value
         cat = sheet.cell(row=i, column=5).value
         brand = sheet.cell(row=i, column=10).value
         img = sheet.cell(row=i, column=2).value.split("'")[1]
         sku =  sheet.cell(row=i, column=6).value
         obem = sheet.cell(row=i, column=12).value
         notes = sheet.cell(row=i, column=16).value
         parfumer = sheet.cell(row=i, column=14).value
         aromat = sheet.cell(row=i, column=13).value

         for c in categories:
             if cat == c['name']:
                 cat = c['id']
                 break

         for c in categories:
             if brand == c['name']:
                 brand = c['id']
                 break

         attributes = []

         notes = str(notes).split(',')

         i_nites = []

         for t in notes:
             if len(t) != 0:
                 i_nites.append(t)

         aromaaat = str(aromat).split(',')

         i_aromaaat = []

         for t in aromaaat:
             if len(t) != 0:
                 i_aromaaat.append(t)

         for elem in wcapi.get("products/attributes").json():
             if elem['name'] == 'Обьём' and str(obem) != '0':
                 attributes.append({'id': elem['id'], 'name': elem['name'], 'options': [str(obem)]})
             if elem['name'] == 'Ноты':
                 attributes.append({'id': elem['id'], 'name': elem['name'], 'options': i_nites})
             if elem['name'] == 'Парфюмер' and str(parfumer) != '0':
                 attributes.append({'id': elem['id'], 'name': elem['name'], 'options': [str(parfumer)]})
             if elem['name'] == 'Ароматы':
                 attributes.append({'id': elem['id'], 'name': elem['name'], 'options': i_aromaaat})



         data = {
             "name": name,
             "type": "simple",
             "regular_price": str(price),
             "description": " ",
             "short_description": " ",
             "sku": str(sku),
             "categories": [
                 {
                     "id": cat
                 },
                 {
                     "id": brand
                 }
             ],
             "images": [
                 {
                     "src": img
                 }
             ],
             'attributes': attributes
         }
         # wcapi.post("products", data).json()
         try:
            wcapi.post("products", data).json()
            time.sleep(1)
         except:
             pass


push_products()